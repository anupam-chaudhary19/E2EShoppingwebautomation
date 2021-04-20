import unittest
import time
from selenium import webdriver
from Pages.Login import login
from Locators.locators import Locators
from Pages.Products import Products
from Pages.About import About
from Pages.Cart import Cart
from Pages.Logout import Logout
from Pages.Allitems import Allitems
from Pages.ResetAppState import Reset
from Locators.Globalvariables import Globalvar
import openpyxl
path = Globalvar.Datasheet
wb_obj = openpyxl.load_workbook(path)
sheet1 = wb_obj.active
un = sheet1.cell(2, 1).value
pwd = sheet1.cell(2, 2).value


class E2ETestsuite(unittest.TestCase):
    @classmethod
    def setUp(cls) -> None:
        cls.driver = webdriver.Chrome(executable_path=Globalvar.Chromepath)
        cls.driver.implicitly_wait(10)
        cls.driver.maximize_window()

    def test_loginvalid_TC001(self):
        driver = self.driver
        driver.get(Globalvar.baseURL)
        log = login(driver)
        log.enterusername(un)
        log.enterpassword(pwd)
        log.Login()
        time.sleep(10)
        pagetitle = driver.title
        if pagetitle == "Swag Labs":
            assert True
        else:
            assert False

    def test_products_addtocartTC002(self):
        driver = self.driver
        driver.get(Globalvar.baseURL)
        log = login(driver)
        log.enterusername(sheet1.cell(5, 1).value)
        log.enterpassword(sheet1.cell(5, 2).value)
        log.Login()
        # User able to select products and proceed to cart items
        prod = Products(driver)
        prod.Backpack_Addtocart()
        prod.Tshirts_Addtocart()
        prod.Click_Add_to_cart()
        time.sleep(5)
        cart = Cart(driver)
        # User Continue to Shop
        cart.Continueshopping_ID()
        # New Item Added in Cart
        prod.Onsie_Addtocart()
        # User fill checkout page details like Fn, Ln etc, Continue with order and Finish
        prod.Click_Add_to_cart()
        cart.Checkout()
        pagetitle1 = driver.title
        if pagetitle1 == "Swag Labs":
            assert True
        else:
            assert False
        time.sleep(5)
        cart.Checkout_Successtext_verify()

    def test_menulinkAbout_TC003(self):
        driver = self.driver
        driver.get(Globalvar.baseURL)
        log = login(driver)
        log.enterusername(sheet1.cell(5, 1).value)
        log.enterpassword(sheet1.cell(5, 2).value)
        log.Login()
        ab = About(driver)
        ab.Click_Access_4items()
        ab.Click_About()
        pagetitle2 = driver.title
        if pagetitle2 == "Cross Browser Testing, Selenium Testing, Mobile Testing | Sauce Labs":
            assert True
        else:
            assert False
        ab.Click_try_it_free()

        ab.enter_workemail(Globalvar.wemail)
        ab.enter_username(Globalvar.uname)
        ab.enter_password(Globalvar.pwd)
        ab.submit()


    def test_menulinkLogout_TC004(self):
        driver = self.driver
        driver.get(Globalvar.baseURL)
        log = login(driver)
        log.enterusername(sheet1.cell(2, 1).value)
        log.enterpassword(sheet1.cell(2, 2).value)
        log.Login()
        lg = Logout(driver)
        lg.Click_Access_4items()
        lg.Click_logout()
        pagetitle2 = driver.title
        if pagetitle2 == "Swag Labs":
            assert True
        else:
            assert False

    def test_menulinkreset_TC005(self):
        driver = self.driver
        driver.get(Globalvar.baseURL)
        log = login(driver)
        log.enterusername(sheet1.cell(2, 1).value)
        log.enterpassword(sheet1.cell(2, 2).value)
        log.Login()
        reset = Reset(driver)
        reset.Click_Access_4items()
        reset.Click_Reset()
        pagetitle2 = driver.title
        if pagetitle2 == "Swag Labs":
            assert True
        else:
            assert False



    def tearDown(self) -> None:
        self.driver.close()
        self.driver.quit()

